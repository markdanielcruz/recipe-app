import mimetypes
mimetypes.init()
mimetypes.add_type(
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.xlsx'
)

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from tempfile import NamedTemporaryFile
from PIL import Image as PILImage
from io import BytesIO
import tempfile
import os
import base64

st.set_page_config(page_title="Servando Recipe Card Generator", layout="wide")

# ── Encode logo ──────────────────────────────────────────────
def get_logo_b64():
    logo_path = os.path.join(os.path.dirname(__file__), "Servando_Branding.jpg")
    try:
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

logo_b64 = get_logo_b64()

# ── CSS ──────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;600;700&family=Inter:wght@300;400;500;600&display=swap');

*, *::before, *::after {{ box-sizing: border-box; }}

html, body, [class*="css"] {{
    font-family: 'Inter', sans-serif;
    background-color: #0E1410;
    color: #D4D0C8;
}}

.stApp {{
    background-color: #0E1410;
}}

/* ── SIDEBAR ── */
[data-testid="stSidebar"] {{
    background-color: #111811 !important;
    border-right: 1px solid #2A3828;
    padding-top: 0 !important;
}}
[data-testid="stSidebar"] > div:first-child {{
    padding-top: 0;
}}

/* ── LOGO PANEL ── */
.logo-panel {{
    background: linear-gradient(175deg, #1A2E1A 0%, #0E1C0E 60%, #080F08 100%);
    padding: 32px 24px 24px 24px;
    border-bottom: 1px solid #2A3828;
    text-align: center;
    position: relative;
    overflow: hidden;
}}
.logo-panel::before {{
    content: '';
    position: absolute;
    inset: 0;
    background: radial-gradient(ellipse at 50% 0%, rgba(74,107,62,0.18) 0%, transparent 70%);
    pointer-events: none;
}}
.logo-panel img {{
    width: 85%;
    max-width: 220px;
    mix-blend-mode: luminosity;
    filter: brightness(0.85) contrast(1.1) sepia(0.2) hue-rotate(60deg) saturate(0.7);
    border-radius: 4px;
}}
.logo-sub {{
    font-family: 'Inter', sans-serif;
    font-size: 0.6rem;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #5A7A52;
    margin-top: 10px;
}}

/* ── SIDEBAR NAV LABELS ── */
.nav-label {{
    font-size: 0.6rem;
    font-weight: 600;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #4A6B3E;
    padding: 20px 0 6px 0;
    border-bottom: 1px solid #1E2E1C;
    margin-bottom: 8px;
}}

/* ── MAIN AREA ── */
.main-header {{
    background: linear-gradient(135deg, #1A2E1A 0%, #152515 100%);
    border: 1px solid #2A3828;
    border-radius: 12px;
    padding: 24px 32px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 20px;
}}
.main-header-text h2 {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 1.6rem;
    color: #8CAF7A;
    margin: 0 0 2px 0;
    font-weight: 600;
    letter-spacing: 1px;
}}
.main-header-text p {{
    font-size: 0.7rem;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    color: #4A6B3E;
    margin: 0;
}}

/* ── SECTION LABEL ── */
.section-label {{
    font-size: 0.62rem;
    font-weight: 600;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #5A7A52;
    margin: 20px 0 10px 0;
    padding-bottom: 6px;
    border-bottom: 1px solid #1E2E1C;
}}

/* ── CARDS ── */
.info-card {{
    background: #111E11;
    border: 1px solid #1E2E1C;
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 12px;
}}

/* ── METRICS ── */
[data-testid="metric-container"] {{
    background: #111E11 !important;
    border: 1px solid #1E2E1C !important;
    border-radius: 10px !important;
    padding: 14px 16px !important;
}}
[data-testid="metric-container"] label {{
    font-size: 0.62rem !important;
    letter-spacing: 2px !important;
    text-transform: uppercase !important;
    color: #5A7A52 !important;
}}
[data-testid="stMetricValue"] {{
    font-family: 'Cormorant Garamond', serif !important;
    font-size: 1.6rem !important;
    font-weight: 600 !important;
    color: #A8C896 !important;
}}

/* ── INPUTS ── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > textarea,
.stSelectbox > div > div {{
    background-color: #111E11 !important;
    border: 1px solid #2A3828 !important;
    border-radius: 8px !important;
    color: #D4D0C8 !important;
    font-size: 0.88rem !important;
}}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > textarea:focus {{
    border-color: #5A7A52 !important;
    box-shadow: 0 0 0 2px rgba(90,122,82,0.2) !important;
}}
label {{
    color: #8A9E84 !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.3px !important;
}}

/* ── BUTTONS ── */
.stButton > button {{
    background-color: #2A3E28 !important;
    color: #A8C896 !important;
    border: 1px solid #3A5238 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.8px !important;
    padding: 9px 18px !important;
    transition: all 0.2s !important;
    width: 100%;
}}
.stButton > button:hover {{
    background-color: #3A5238 !important;
    border-color: #5A7A52 !important;
    color: #C8DCC0 !important;
}}

/* Primary generate button */
.generate-btn .stButton > button {{
    background: linear-gradient(135deg, #3A5238 0%, #2A3E28 100%) !important;
    border: 1px solid #5A7A52 !important;
    color: #C8DCC0 !important;
    font-size: 0.9rem !important;
    padding: 14px 24px !important;
    border-radius: 10px !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
}}

/* Download button */
[data-testid="stDownloadButton"] > button {{
    background: linear-gradient(135deg, #3A5238, #2A3E28) !important;
    color: #C8DCC0 !important;
    border: 1px solid #5A7A52 !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    width: 100% !important;
    padding: 14px !important;
}}

/* ── EXPANDER ── */
.streamlit-expanderHeader {{
    background-color: #111E11 !important;
    border: 1px solid #1E2E1C !important;
    border-radius: 8px !important;
    color: #A8C896 !important;
    font-size: 0.86rem !important;
    font-weight: 500 !important;
}}
.streamlit-expanderContent {{
    background-color: #0E1810 !important;
    border: 1px solid #1E2E1C !important;
    border-top: none !important;
    border-radius: 0 0 8px 8px !important;
}}

/* ── DATAFRAME ── */
[data-testid="stDataFrame"] {{
    border: 1px solid #1E2E1C !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}}

/* ── DIVIDER ── */
hr {{ border-color: #1E2E1C !important; }}

/* ── SUCCESS/ERROR ── */
[data-testid="stAlert"] {{ border-radius: 8px !important; }}
.stSuccess {{ background-color: #1A2E1A !important; border-color: #3A5238 !important; }}
.stError {{ background-color: #2E1A1A !important; }}

/* ── FILE UPLOADER ── */
[data-testid="stFileUploader"] {{
    border: 1px dashed #2A3828 !important;
    border-radius: 10px !important;
    background: #111E11 !important;
    padding: 12px !important;
}}

/* ── SUBHEADER ── */
h3 {{
    font-family: 'Cormorant Garamond', serif !important;
    color: #A8C896 !important;
    font-size: 1.3rem !important;
}}

/* ── TOTAL COST BADGE ── */
.total-badge {{
    background: linear-gradient(135deg, #1A2E1A, #111E11);
    border: 1px solid #3A5238;
    border-radius: 10px;
    padding: 16px 24px;
    text-align: center;
    margin: 12px 0;
}}
.total-badge .label {{
    font-size: 0.62rem;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #5A7A52;
}}
.total-badge .value {{
    font-family: 'Inter', sans-serif;
    font-size: 1.8rem;
    font-weight: 600;
    color: #A8C896;
    display: block;
    margin-top: 4px;
}}
</style>
""", unsafe_allow_html=True)

# ── SIDEBAR ─────────────────────────────────────────────────
with st.sidebar:
    # Logo panel
    if logo_b64:
        st.markdown(f"""
        <div class="logo-panel">
            <img src="data:image/jpeg;base64,{logo_b64}" alt="Servando Logo"/>
            <div class="logo-sub">Recipe Card System</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="logo-panel">
            <div style="font-family:'Cormorant Garamond',serif;font-size:1.8rem;color:#8CAF7A;font-weight:700;">SERVANDO</div>
            <div class="logo-sub">Recipe Card System</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('<div class="nav-label">📌 Recipe Details</div>', unsafe_allow_html=True)
    recipe_name = st.text_input("Recipe Name")
    category    = st.text_input("Category")

    st.markdown('<div class="nav-label">⚖️ Yield & Serving</div>', unsafe_allow_html=True)
    total_yield  = st.number_input("Total Recipe Yield", min_value=0.0)
    serving_size = st.number_input("Serving Size",        min_value=0.0)
    servings = total_yield / serving_size if serving_size > 0 else 0
    st.metric("No. of Servings", f"{servings:.0f}")

    st.markdown('<div class="nav-label">✍️ Sign-Off</div>', unsafe_allow_html=True)
    prepared_by = st.text_input("Prepared By")
    checked_by  = st.text_input("Checked By")

    st.markdown('<div class="nav-label">📷 Photos</div>', unsafe_allow_html=True)
    images = st.file_uploader("Upload Photos", type=["png","jpg","jpeg"], accept_multiple_files=True)

# ── LOAD COSTS ───────────────────────────────────────────────
cost_df = pd.read_excel("costs.xlsx")
cost_df.columns = cost_df.columns.str.strip()
ingredients = cost_df.iloc[:, 0].dropna().tolist()

if "items" not in st.session_state:
    st.session_state["items"] = []

# ── MAIN HEADER ──────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <div class="main-header-text">
        <h2>Recipe Card Generator</h2>
        <p>Costing &nbsp;·&nbsp; Standardization &nbsp;·&nbsp; Production</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ── TWO COLUMN LAYOUT ────────────────────────────────────────
left, right = st.columns([1.1, 1], gap="large")

# ════════════════ LEFT COLUMN ════════════════
with left:
    st.markdown('<div class="section-label">➕ Add Ingredient</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        ingredient = st.selectbox("Ingredient", ingredients)
    with col2:
        qty = st.number_input("Quantity", min_value=0.0, key="qty_input")

    notes_input = st.text_input("Notes", placeholder="e.g. sifted, room temp, finely chopped…")

    if st.button("＋ Add Ingredient"):
        try:
            row = cost_df[cost_df.iloc[:, 0] == ingredient].iloc[0]
            unit_cost = float(row.iloc[1])
            uom       = str(row.iloc[2])
            packaging = 1000 if uom.lower() in ["g", "ml"] else 1
            st.session_state["items"].append({
                "ingredient": ingredient,
                "qty":        qty,
                "packaging":  packaging,
                "uom":        uom,
                "unit_cost":  unit_cost,
                "notes":      notes_input
            })
            st.success(f"Added: {ingredient}")
        except:
            st.error("Error adding ingredient. Check costs.xlsx format.")

    st.markdown('<div class="section-label">📋 Ingredients List</div>', unsafe_allow_html=True)

    delete_index = None
    for i, item in enumerate(st.session_state["items"]):
        with st.expander(f"{item['ingredient']}  —  {item['qty']} {item['uom']}"):
            c1, c2 = st.columns(2)
            with c1:
                item["qty"]       = st.number_input("Qty",       value=float(item["qty"]),       key=f"qty_{i}")
                item["packaging"] = st.number_input("Packaging", value=float(item["packaging"]), key=f"pack_{i}")
            with c2:
                item["uom"]       = st.text_input("UOM",       value=item["uom"],              key=f"uom_{i}")
                item["unit_cost"] = st.number_input("Unit Cost", value=float(item["unit_cost"]), key=f"cost_{i}")
            item["notes"] = st.text_input("Notes", value=item.get("notes",""), key=f"notes_{i}", placeholder="e.g. sifted…")
            if st.button(f"Remove", key=f"delete_{i}"):
                delete_index = i

    if delete_index is not None:
        st.session_state["items"].pop(delete_index)
        st.rerun()

    c_undo, c_clear = st.columns(2)
    with c_undo:
        if st.button("↩ Undo Last") and st.session_state["items"]:
            removed = st.session_state["items"].pop()
            st.toast(f"Removed: {removed['ingredient']}")
            st.rerun()
    with c_clear:
        if st.button("✕ Clear All"):
            st.session_state["items"] = []
            st.rerun()

    st.markdown('<div class="section-label">🧑‍🍳 Procedure</div>', unsafe_allow_html=True)

    if "procedure_steps" not in st.session_state:
        st.session_state["procedure_steps"] = [""]

    for idx in range(len(st.session_state["procedure_steps"])):
        st.session_state["procedure_steps"][idx] = st.text_input(
            f"Step {idx + 1}",
            value=st.session_state["procedure_steps"][idx],
            key=f"step_{idx}",
            placeholder="Describe this step…"
        )

    cp1, cp2 = st.columns(2)
    with cp1:
        if st.button("＋ Add Step"):
            st.session_state["procedure_steps"].append("")
            st.rerun()
    with cp2:
        if st.button("− Remove Last Step") and len(st.session_state["procedure_steps"]) > 1:
            st.session_state["procedure_steps"].pop()
            st.rerun()

    # Build procedure string for Excel export
    procedure = "\n".join(st.session_state["procedure_steps"])

# ════════════════ RIGHT COLUMN ════════════════
with right:
    total = 0.0
    srp   = 0.0

    if st.session_state["items"]:
        st.markdown('<div class="section-label">📊 Cost Breakdown</div>', unsafe_allow_html=True)

        df = pd.DataFrame(st.session_state["items"])
        df["Total Cost"] = df["qty"] * df["unit_cost"]
        display_df = df[["ingredient","qty","uom","unit_cost","Total Cost"]].copy()
        display_df.columns = ["Ingredient","Qty","UOM","Unit Cost","Total Cost"]
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        total = df["Total Cost"].sum()
        st.markdown(f"""
        <div class="total-badge">
            <div class="label">Total Recipe Cost</div>
            <span class="value">₱{total:,.2f}</span>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="section-label">💰 Pricing</div>', unsafe_allow_html=True)
        srp = st.number_input("Selling Price (SRP)", min_value=0.0)

        if srp > 0:
            fc_pct = (total / srp) * 100
            m1, m2 = st.columns(2)
            with m1:
                st.metric("Food Cost %", f"{fc_pct:.1f}%")
            with m2:
                st.metric("Gross Profit", f"₱{srp - total:,.2f}")

    else:
        st.markdown("""
        <div style="text-align:center;padding:60px 20px;color:#3A5238;border:1px dashed #1E2E1C;border-radius:12px;margin-top:40px;">
            <div style="font-size:2rem;margin-bottom:12px;">🍽</div>
            <div style="font-family:'Cormorant Garamond',serif;font-size:1.1rem;color:#5A7A52;">No ingredients added yet</div>
            <div style="font-size:0.75rem;letter-spacing:1px;margin-top:6px;color:#3A5238;">Use the form on the left to begin</div>
        </div>
        """, unsafe_allow_html=True)

    # ── GENERATE ──────────────────────────────────────────
    st.markdown('<div class="section-label">📥 Export</div>', unsafe_allow_html=True)

    if st.button("⬇  Generate Recipe Card"):
        try:
            wb = load_workbook("template.xlsx")
            ws = wb.active

            ws["A3"]  = recipe_name
            ws["A6"]  = category
            ws["A55"] = recipe_name
            ws["A58"] = category
            ws["A8"]  = total_yield
            ws["C8"]  = serving_size
            ws["G48"] = srp
            ws["H47"] = prepared_by
            ws["H51"] = checked_by

            start_row = 13
            for i, item in enumerate(st.session_state["items"]):
                r = start_row + i
                ws[f"A{r}"] = item["ingredient"]
                ws[f"B{r}"] = item["qty"]
                ws[f"C{r}"] = item["packaging"]
                ws[f"D{r}"] = item["uom"]
                ws[f"F{r}"] = item["unit_cost"]
                ws[f"H{r}"] = item.get("notes", "")

            for r in range(start_row + len(st.session_state["items"]), 41):
                for col in ["A","B","C","D","E","F","G","H"]:
                    ws[f"{col}{r}"] = None

            row_cursor = 62
            for i, line in enumerate(st.session_state["procedure_steps"], start=1):
                if line.strip():
                    ws[f"A{row_cursor}"] = f"Step {i}: {line}"
                    row_cursor += 1

            START_ROW = 66
            COLS      = ["A","D","G"]
            IMG_W, IMG_H = 240, 160
            row_pos   = START_ROW
            col_index = 0
            temp_images = []

            for img_file in images or []:
                try:
                    img_file.seek(0)
                    pil_img = PILImage.open(img_file).convert("RGB")
                    buf = BytesIO()
                    pil_img.save(buf, format="PNG")
                    buf.seek(0)
                    with NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                        tmp.write(buf.read())
                        tp = tmp.name
                    temp_images.append(tp)
                    img = XLImage(tp)
                    img.width, img.height = IMG_W, IMG_H
                    ws.add_image(img, f"{COLS[col_index]}{row_pos}")
                    col_index += 1
                    if col_index == 3:
                        col_index = 0
                        row_pos  += 16
                except:
                    pass

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                temp_path = tmp.name
            wb.save(temp_path)
            for p in temp_images:
                try: os.remove(p)
                except: pass

            with open(temp_path, "rb") as f:
                file_data = f.read()
            os.remove(temp_path)

            file_name = f"{recipe_name.strip().replace(' ','_')}.xlsx" if recipe_name else "recipe.xlsx"

            st.download_button(
                label="📥  Download Recipe Card",
                data=file_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Failed to generate: {e}")

# ── FOOTER ───────────────────────────────────────────────────
st.markdown("""
<div style="text-align:center;padding:32px 0 16px 0;">
    <span style="font-size:0.68rem;letter-spacing:2px;text-transform:uppercase;color:#2A3E28;">
        developed by Dong Cruz
    </span>
</div>
""", unsafe_allow_html=True)